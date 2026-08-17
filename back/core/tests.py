from datetime import date, datetime, time, timedelta
from io import BytesIO

from accounts.models import User
from django.db import IntegrityError, transaction
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from .models import Attendance, QRToken, SystemSettings
from .views import (
    DETAIL_ROW_ABSENT_DAYS,
    DETAIL_ROW_DUTY_TYPE,
    DETAIL_ROW_EARLY_LEAVE_MINUTES,
    DETAIL_ROW_LATE_MINUTES,
    DETAIL_ROW_NAME,
    DETAIL_ROW_PHONE,
    DETAIL_ROW_PRESENT_DAYS,
    DETAIL_ROW_TYPE,
    _count_working_days_between,
    _working_days_in_range,
)


class GenerateQRTokenTests(APITestCase):
    url = "/api/qr/generate/"

    def setUp(self):
        self.entry = User.objects.create_user(username="entry1", password="x", is_entry=True)
        self.employee = User.objects.create_user(username="emp1", password="x", is_employee=True)

    def test_entry_can_generate(self):
        self.client.force_authenticate(self.entry)
        response = self.client.post(self.url, {}, format="json")
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertIn("token", response.data)

    def test_employee_cannot_generate(self):
        self.client.force_authenticate(self.employee)
        response = self.client.post(self.url, {}, format="json")
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_uses_system_settings_lifetime(self):
        settings_obj = SystemSettings.get_solo()
        settings_obj.qr_token_lifetime_seconds = 30
        settings_obj.save()

        self.client.force_authenticate(self.entry)
        response = self.client.post(self.url, {}, format="json")
        token = QRToken.objects.get(token=response.data["token"])
        delta = (token.expires_at - token.created_at).total_seconds()
        self.assertAlmostEqual(delta, 30, delta=1)


class ValidateQRTests(APITestCase):
    url = "/api/qr/validate/"

    def setUp(self):
        self.entry = User.objects.create_user(username="entry1", password="x", is_entry=True)
        self.employee = User.objects.create_user(username="emp1", password="x", is_employee=True)
        self.client.force_authenticate(self.employee)

    def _make_token(self, expired=False):
        return QRToken.objects.create(
            token="tok123",
            generated_by=self.entry,
            expires_at=timezone.now() + (timedelta(seconds=-1) if expired else timedelta(seconds=15)),
        )

    def test_predicts_check_in_when_no_open_session(self):
        self._make_token()
        response = self.client.post(self.url, {"token": "tok123"}, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data["valid"])
        self.assertEqual(response.data["action"], "check_in")

    def test_predicts_check_out_when_open_session_exists(self):
        Attendance.objects.create(user=self.employee, date=timezone.localdate(), check_in=timezone.now())
        self._make_token()
        response = self.client.post(self.url, {"token": "tok123"}, format="json")
        self.assertEqual(response.data["action"], "check_out")

    def test_nonexistent_token(self):
        response = self.client.post(self.url, {"token": "nope"}, format="json")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_expired_token(self):
        self._make_token(expired=True)
        response = self.client.post(self.url, {"token": "tok123"}, format="json")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)


class RecordAttendanceTests(APITestCase):
    url = "/api/attendance/record/"

    def setUp(self):
        self.entry = User.objects.create_user(username="entry1", password="x", is_entry=True)
        self.employee = User.objects.create_user(username="emp1", password="x", is_employee=True)
        self.client.force_authenticate(self.employee)
        # يُعطَّل الحد الأدنى الزمني هنا لعزل اختبار منطق الحالة عن اختبار الفجوة الزمنية (له اختبار مخصص أدناه)
        settings_obj = SystemSettings.get_solo()
        settings_obj.min_session_duration_seconds = 0
        settings_obj.save()

    def _token(self):
        return QRToken.objects.create(
            token=f"tok-{timezone.now().timestamp()}",
            generated_by=self.entry,
            expires_at=timezone.now() + timedelta(seconds=15),
        )

    def test_first_scan_checks_in(self):
        response = self.client.post(self.url, {"token": self._token().token}, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsNotNone(response.data["check_in"])
        self.assertIsNone(response.data["check_out"])

    def test_second_scan_checks_out(self):
        self.client.post(self.url, {"token": self._token().token}, format="json")
        response = self.client.post(self.url, {"token": self._token().token}, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsNotNone(response.data["check_out"])

    def test_third_scan_same_day_opens_new_session(self):
        self.client.post(self.url, {"token": self._token().token}, format="json")  # check-in
        self.client.post(self.url, {"token": self._token().token}, format="json")  # check-out
        response = self.client.post(self.url, {"token": self._token().token}, format="json")  # check-in again
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsNone(response.data["check_out"])
        self.assertEqual(
            Attendance.objects.filter(user=self.employee, date=timezone.localdate()).count(), 2
        )

    def test_two_open_sessions_blocked_at_database_level(self):
        Attendance.objects.create(user=self.employee, date=timezone.localdate(), check_in=timezone.now())
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Attendance.objects.create(
                    user=self.employee, date=timezone.localdate(), check_in=timezone.now()
                )

    def test_entry_role_cannot_record(self):
        self.client.force_authenticate(self.entry)
        token = self._token()
        response = self.client.post(self.url, {"token": token.token}, format="json")
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


class MinSessionDurationTests(APITestCase):
    """يتحقق من الفجوة الزمنية الدنيا التي تمنع مسح واحد سريع يسجّل دخولًا وخروجًا فوريين."""

    url = "/api/attendance/record/"

    def setUp(self):
        self.entry = User.objects.create_user(username="entry1", password="x", is_entry=True)
        self.employee = User.objects.create_user(username="emp1", password="x", is_employee=True)
        self.client.force_authenticate(self.employee)
        settings_obj = SystemSettings.get_solo()
        settings_obj.min_session_duration_seconds = 300
        settings_obj.save()

    def _token(self):
        return QRToken.objects.create(
            token=f"tok-{timezone.now().timestamp()}",
            generated_by=self.entry,
            expires_at=timezone.now() + timedelta(seconds=15),
        )

    def test_checkout_rejected_before_minimum_duration(self):
        self.client.post(self.url, {"token": self._token().token}, format="json")
        response = self.client.post(self.url, {"token": self._token().token}, format="json")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("wait", response.data["detail"].lower())

    def test_checkout_allowed_after_minimum_duration(self):
        self.client.post(self.url, {"token": self._token().token}, format="json")
        attendance = Attendance.objects.get(user=self.employee, date=timezone.localdate())
        attendance.check_in = timezone.now() - timedelta(seconds=301)
        attendance.save()

        response = self.client.post(self.url, {"token": self._token().token}, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsNotNone(response.data["check_out"])


class MyAttendanceTests(APITestCase):
    url = "/api/attendance/my/"

    def setUp(self):
        self.employee = User.objects.create_user(username="emp1", password="x", is_employee=True)
        self.other = User.objects.create_user(username="emp2", password="x", is_employee=True)
        self.client.force_authenticate(self.employee)

    def test_shows_only_current_month(self):
        today = timezone.localdate()
        last_month = today.replace(day=1) - timedelta(days=1)
        Attendance.objects.create(user=self.employee, date=today, check_in=timezone.now())
        Attendance.objects.create(
            user=self.employee, date=last_month, check_in=timezone.now(), check_out=timezone.now()
        )
        response = self.client.get(self.url)
        self.assertEqual(response.data["count"], 1)

    def test_isolated_per_user(self):
        Attendance.objects.create(user=self.other, date=timezone.localdate(), check_in=timezone.now())
        response = self.client.get(self.url)
        self.assertEqual(response.data["count"], 0)


class EmployeeAdminViewTests(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(username="admin", password="x", is_admin=True)
        self.employee = User.objects.create_user(
            username="emp1", password="x", is_employee=True, first_name="Test", last_name="Employee"
        )
        self.entry = User.objects.create_user(username="entry1", password="x", is_entry=True)
        self.client.force_authenticate(self.admin)

    def test_list_filter_by_is_employee(self):
        response = self.client.get("/api/admin/employees/", {"is_employee": "true"})
        usernames = [u["username"] for u in response.data["results"]]
        self.assertIn("emp1", usernames)
        self.assertNotIn("entry1", usernames)

    def test_list_is_paginated(self):
        for i in range(15):
            User.objects.create_user(username=f"bulk{i}", password="x", is_employee=True)
        response = self.client.get("/api/admin/employees/", {"is_employee": "true"})
        self.assertEqual(len(response.data["results"]), 10)
        self.assertEqual(response.data["count"], 16)
        self.assertIsNotNone(response.data["next"])

    def test_detail_includes_attendance(self):
        Attendance.objects.create(user=self.employee, date=timezone.localdate(), check_in=timezone.now())
        response = self.client.get(f"/api/admin/employees/{self.employee.pk}/")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["attendance"]), 1)

    def test_detail_includes_month_stats(self):
        settings_obj = SystemSettings.get_solo()
        settings_obj.work_start_time = time(9, 0)
        settings_obj.work_end_time = time(17, 0)
        settings_obj.save()

        Attendance.objects.create(
            user=self.employee,
            date=date(2024, 1, 8),
            check_in=timezone.make_aware(datetime(2024, 1, 8, 9, 15)),
            check_out=timezone.make_aware(datetime(2024, 1, 8, 16, 30)),
        )

        response = self.client.get(f"/api/admin/employees/{self.employee.pk}/", {"year": 2024, "month": 1})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        working_days = _working_days_in_range(2024, 1, 31)
        self.assertEqual(response.data["present_days"], 1)
        self.assertEqual(response.data["absent_days"], working_days - 1)
        self.assertEqual(response.data["late_minutes"], 15)
        self.assertEqual(response.data["early_leave_minutes"], 30)

    def test_detail_404_for_non_employee(self):
        response = self.client.get(f"/api/admin/employees/{self.entry.pk}/")
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_non_admin_forbidden(self):
        self.client.force_authenticate(self.employee)
        response = self.client.get("/api/admin/employees/")
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


class EntryUserListViewTests(APITestCase):
    url = "/api/admin/entry-users/"

    def setUp(self):
        self.admin = User.objects.create_superuser(username="admin", password="x", is_admin=True)
        self.entry = User.objects.create_user(username="entry1", password="x", is_entry=True)
        self.employee = User.objects.create_user(username="emp1", password="x", is_employee=True)
        self.client.force_authenticate(self.admin)

    def test_list_shows_entry_users_only(self):
        response = self.client.get(self.url)
        usernames = [u["username"] for u in response.data["results"]]
        self.assertIn("entry1", usernames)
        self.assertNotIn("emp1", usernames)

    def test_search_by_username(self):
        response = self.client.get(self.url, {"search": "entry1"})
        usernames = [u["username"] for u in response.data["results"]]
        self.assertEqual(usernames, ["entry1"])

    def test_non_admin_forbidden(self):
        self.client.force_authenticate(self.entry)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


class EmployeeExportAttendanceTests(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(username="admin", password="x", is_admin=True)
        self.employee = User.objects.create_user(username="emp1", password="x", is_employee=True)
        self.entry = User.objects.create_user(username="entry1", password="x", is_entry=True)
        self.client.force_authenticate(self.admin)
        self.url = f"/api/admin/employees/{self.employee.pk}/export/"

    def test_export_succeeds(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_export_only_contains_this_employee(self):
        other = User.objects.create_user(username="emp2", password="x", is_employee=True)
        Attendance.objects.create(user=self.employee, date=timezone.localdate(), check_in=timezone.now())
        Attendance.objects.create(user=other, date=timezone.localdate(), check_in=timezone.now())

        response = self.client.get(self.url)
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        self.assertEqual(rows[0], ("التاريخ", "وقت الحضور", "وقت الانصراف"))
        self.assertEqual(len(rows), 2)

    def test_404_for_non_employee(self):
        response = self.client.get(f"/api/admin/employees/{self.entry.pk}/export/")
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_non_admin_forbidden(self):
        self.client.force_authenticate(self.employee)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_invalid_start_date_returns_400_not_500(self):
        response = self.client.get(self.url, {"start_date": "not-a-date"})
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_start_date_after_end_date_rejected(self):
        response = self.client.get(
            self.url, {"start_date": "2026-08-20", "end_date": "2026-08-01"}
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_only_records_within_date_range_are_included(self):
        Attendance.objects.create(
            user=self.employee, date=date(2026, 1, 5), check_in=timezone.now()
        )
        Attendance.objects.create(
            user=self.employee, date=date(2026, 1, 15), check_in=timezone.now()
        )
        Attendance.objects.create(
            user=self.employee, date=date(2026, 2, 1), check_in=timezone.now()
        )

        response = self.client.get(
            self.url, {"start_date": "2026-01-01", "end_date": "2026-01-31"}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(len(rows), 3)  # عنوان + سطران ضمن يناير فقط
        self.assertEqual(rows[1][0], "2026-01-05")
        self.assertEqual(rows[2][0], "2026-01-15")


class WorkingDaysCalculationTests(APITestCase):
    def test_excludes_saturdays_and_sundays(self):
        # يناير 2024: يبدأ الاثنين 1/1، 31 يومًا، منها 8 أيام عطلة أسبوعية (سبت/أحد)
        self.assertEqual(_working_days_in_range(2024, 1, 31), 23)

    def test_partial_month_up_to_a_given_day(self):
        # أول أسبوع من يناير 2024 (1 اثنين إلى 7 أحد): 5 أيام عمل + عطلة نهاية الأسبوع
        self.assertEqual(_working_days_in_range(2024, 1, 7), 5)


class DateRangeWorkingDaysTests(APITestCase):
    def test_excludes_weekends_within_a_free_range(self):
        # 2024-01-01 (اثنين) إلى 2024-01-07 (أحد): 5 أيام عمل
        self.assertEqual(_count_working_days_between(date(2024, 1, 1), date(2024, 1, 7)), 5)

    def test_caps_at_today_when_range_extends_into_the_future(self):
        today = timezone.localdate()
        future_end = today + timedelta(days=30)
        self.assertEqual(
            _count_working_days_between(today, future_end),
            _count_working_days_between(today, today),
        )

    def test_returns_zero_for_a_fully_future_range(self):
        today = timezone.localdate()
        start = today + timedelta(days=5)
        end = today + timedelta(days=10)
        self.assertEqual(_count_working_days_between(start, end), 0)


class MonthlyAttendanceSummaryExportTests(APITestCase):
    url = "/api/admin/attendance/summary-export/"

    def setUp(self):
        self.admin = User.objects.create_superuser(username="admin", password="x", is_admin=True)
        self.employee = User.objects.create_user(
            username="emp1",
            password="x",
            is_employee=True,
            first_name="Test",
            last_name="Employee",
            is_regular=True,
        )
        self.client.force_authenticate(self.admin)
        settings_obj = SystemSettings.get_solo()
        settings_obj.work_start_time = time(9, 0)
        settings_obj.work_end_time = time(17, 0)
        settings_obj.save()

    def test_non_admin_forbidden(self):
        self.client.force_authenticate(self.employee)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_invalid_start_date_returns_400(self):
        response = self.client.get(self.url, {"start_date": "abc"})
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_start_date_after_end_date_rejected(self):
        response = self.client.get(
            self.url, {"start_date": "2024-01-31", "end_date": "2024-01-01"}
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_report_contents_for_past_month(self):
        # الاثنين: حضور متأخر 15 دقيقة، وانصراف بالوقت تمامًا
        Attendance.objects.create(
            user=self.employee,
            date=date(2024, 1, 8),
            check_in=timezone.make_aware(datetime(2024, 1, 8, 9, 15)),
            check_out=timezone.make_aware(datetime(2024, 1, 8, 17, 0)),
        )
        # الثلاثاء: حضور بالوقت، وانصراف مبكر 30 دقيقة
        Attendance.objects.create(
            user=self.employee,
            date=date(2024, 1, 9),
            check_in=timezone.make_aware(datetime(2024, 1, 9, 8, 55)),
            check_out=timezone.make_aware(datetime(2024, 1, 9, 16, 30)),
        )

        response = self.client.get(
            self.url, {"start_date": "2024-01-01", "end_date": "2024-01-31"}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        workbook = load_workbook(BytesIO(response.content))
        summary_sheet = workbook.active
        rows = list(summary_sheet.iter_rows(values_only=True))
        self.assertEqual(
            rows[0],
            (
                "اسم الموظف",
                "نوع الموظف",
                "نوع الدوام",
                "أيام الدوام",
                "أيام الغياب",
                "دقائق التأخير",
                "دقائق الانصراف المبكر",
            ),
        )

        # القيم الفعلية تُخزَّن في ورقة الموظف الفرعية (مصدر الحقيقة الوحيد)، وورقة الملخص
        # تعرضها فقط عبر صيغ ربط (انظر الاختبار التالي).
        detail_sheet = workbook["Test Employee"]
        working_days = _count_working_days_between(date(2024, 1, 1), date(2024, 1, 31))
        self.assertEqual(detail_sheet.cell(row=DETAIL_ROW_TYPE, column=2).value, "غير محدد")
        self.assertEqual(detail_sheet.cell(row=DETAIL_ROW_DUTY_TYPE, column=2).value, "دوام كامل")
        self.assertEqual(detail_sheet.cell(row=DETAIL_ROW_PRESENT_DAYS, column=2).value, 2)
        self.assertEqual(
            detail_sheet.cell(row=DETAIL_ROW_ABSENT_DAYS, column=2).value, working_days - 2
        )
        self.assertEqual(detail_sheet.cell(row=DETAIL_ROW_LATE_MINUTES, column=2).value, 15)
        self.assertEqual(detail_sheet.cell(row=DETAIL_ROW_EARLY_LEAVE_MINUTES, column=2).value, 30)

    def test_summary_sheet_links_to_employee_detail_sheet(self):
        self.employee.type = 3
        self.employee.phone = "0912345678"
        self.employee.save()
        Attendance.objects.create(
            user=self.employee,
            date=date(2024, 1, 8),
            check_in=timezone.make_aware(datetime(2024, 1, 8, 9, 15)),
            check_out=timezone.make_aware(datetime(2024, 1, 8, 17, 0)),
        )

        response = self.client.get(
            self.url, {"start_date": "2024-01-01", "end_date": "2024-01-31"}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        workbook = load_workbook(BytesIO(response.content))
        summary_sheet = workbook.active

        # صف الموظف بورقة الملخص يجب أن يكون كله صيغ ربط بورقته الفرعية، وليس قيمًا ثابتة —
        # هذا ما يجعل تعديل القيمة بالورقة الفرعية ينعكس تلقائيًا هنا عند فتح الملف بإكسل.
        name_cell = summary_sheet["A2"]
        self.assertEqual(name_cell.value, f"='Test Employee'!A{DETAIL_ROW_NAME}")
        self.assertIsNotNone(name_cell.hyperlink)
        self.assertEqual(summary_sheet["B2"].value, f"='Test Employee'!B{DETAIL_ROW_TYPE}")
        self.assertEqual(summary_sheet["C2"].value, f"='Test Employee'!B{DETAIL_ROW_DUTY_TYPE}")
        self.assertEqual(summary_sheet["D2"].value, f"='Test Employee'!B{DETAIL_ROW_PRESENT_DAYS}")
        self.assertEqual(summary_sheet["E2"].value, f"='Test Employee'!B{DETAIL_ROW_ABSENT_DAYS}")
        self.assertEqual(summary_sheet["F2"].value, f"='Test Employee'!B{DETAIL_ROW_LATE_MINUTES}")
        self.assertEqual(
            summary_sheet["G2"].value, f"='Test Employee'!B{DETAIL_ROW_EARLY_LEAVE_MINUTES}"
        )

        self.assertIn("Test Employee", workbook.sheetnames)
        detail_sheet = workbook["Test Employee"]
        self.assertEqual(detail_sheet.cell(row=DETAIL_ROW_NAME, column=1).value, "Test Employee")
        self.assertEqual(detail_sheet.cell(row=DETAIL_ROW_PHONE, column=2).value, "0912345678")
        self.assertEqual(detail_sheet.cell(row=DETAIL_ROW_TYPE, column=2).value, 3)
        detail_rows = list(detail_sheet.iter_rows(values_only=True))
        self.assertIn(("التاريخ", "وقت الحضور", "وقت الانصراف"), detail_rows)
        self.assertIn(("2024-01-08", "09:15:00", "17:00:00"), detail_rows)

    def test_duplicate_employee_names_get_distinct_detail_sheets(self):
        User.objects.create_user(
            username="emp2",
            password="x",
            is_employee=True,
            first_name="Test",
            last_name="Employee",
        )

        response = self.client.get(
            self.url, {"start_date": "2024-01-01", "end_date": "2024-01-31"}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        workbook = load_workbook(BytesIO(response.content))
        sheet_names = [name for name in workbook.sheetnames if name != workbook.sheetnames[0]]
        self.assertEqual(len(sheet_names), len(set(sheet_names)))
        self.assertEqual(len(sheet_names), 2)


class SystemSettingsTests(APITestCase):
    url = "/api/admin/settings/"

    def setUp(self):
        self.admin = User.objects.create_superuser(username="admin", password="x", is_admin=True)
        self.employee = User.objects.create_user(username="emp1", password="x", is_employee=True)

    def test_default_value(self):
        self.client.force_authenticate(self.admin)
        response = self.client.get(self.url)
        self.assertEqual(response.data["qr_token_lifetime_seconds"], 15)
        self.assertEqual(response.data["min_session_duration_seconds"], 60)
        self.assertEqual(response.data["work_start_time"], "09:00:00")
        self.assertEqual(response.data["work_end_time"], "17:00:00")

    def test_admin_can_update(self):
        self.client.force_authenticate(self.admin)
        response = self.client.patch(self.url, {"qr_token_lifetime_seconds": 30}, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(SystemSettings.get_solo().qr_token_lifetime_seconds, 30)

    def test_out_of_range_rejected(self):
        self.client.force_authenticate(self.admin)
        response = self.client.patch(self.url, {"qr_token_lifetime_seconds": 3000}, format="json")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_non_admin_forbidden(self):
        self.client.force_authenticate(self.employee)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
