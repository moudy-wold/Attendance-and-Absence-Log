import calendar
import re
import secrets
from datetime import date, timedelta

from accounts.models import User
from accounts.permissions import IsAdminUser, IsEmployeeUser, IsEntryUser
from accounts.serializers import UserSerializer
from django.db import IntegrityError, transaction
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import OpenApiParameter, extend_schema
from openpyxl import Workbook
from openpyxl.styles import Font
from rest_framework import filters as drf_filters
from rest_framework import generics, status
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Action, Attendance, QRToken, SystemSettings
from .serializers import (
    AdminStatsOverviewSerializer,
    AttendanceSerializer,
    DateRangeQuerySerializer,
    EmployeeAttendanceSerializer,
    EmployeeStatsSerializer,
    QRTokenInputSerializer,
    QRTokenSerializer,
    SystemSettingsSerializer,
    ValidateQRResponseSerializer,
    YearMonthQuerySerializer,
)


def _resolve_year_month(request) -> tuple[int, int]:
    query = YearMonthQuerySerializer(data=request.query_params)
    query.is_valid(raise_exception=True)
    today = timezone.localdate()
    return (
        query.validated_data.get("year", today.year),
        query.validated_data.get("month", today.month),
    )

YEAR_MONTH_PARAMETERS = [
    OpenApiParameter("year", OpenApiTypes.INT, OpenApiParameter.QUERY, required=False),
    OpenApiParameter("month", OpenApiTypes.INT, OpenApiParameter.QUERY, required=False),
]


def _resolve_date_range(request) -> tuple[date, date]:
    """تُرجع (تاريخ البداية، تاريخ النهاية)، وتفترض افتراضيًا الشهر الحالي (من أول يوم فيه حتى
    اليوم) عند غياب أحد الطرفين أو كليهما."""
    query = DateRangeQuerySerializer(data=request.query_params)
    query.is_valid(raise_exception=True)
    today = timezone.localdate()
    start_date = query.validated_data.get("start_date") or today.replace(day=1)
    end_date = query.validated_data.get("end_date") or today
    if start_date > end_date:
        raise ValidationError({"detail": "start_date must be before or equal to end_date."})
    return start_date, end_date

DATE_RANGE_PARAMETERS = [
    OpenApiParameter("start_date", OpenApiTypes.DATE, OpenApiParameter.QUERY, required=False),
    OpenApiParameter("end_date", OpenApiTypes.DATE, OpenApiParameter.QUERY, required=False),
]


class GenerateQRTokenView(APIView):
    """يولّد رمز QR عامًا — لا يحمل نوع إجراء، يُحدَّد ذلك تلقائيًا وقت المسح حسب حالة كل موظف."""

    permission_classes = [IsEntryUser]

    @extend_schema(request=None, responses={201: QRTokenSerializer})
    def post(self, request):
        lifetime_seconds = SystemSettings.get_solo().qr_token_lifetime_seconds
        qr_token = QRToken.objects.create(
            token=secrets.token_urlsafe(24),
            generated_by=request.user,
            expires_at=timezone.now() + timedelta(seconds=lifetime_seconds),
        )
        return Response(
            QRTokenSerializer(qr_token).data, status=status.HTTP_201_CREATED
        )


def _has_open_session(user) -> bool:
    return Attendance.objects.filter(
        user=user, date=timezone.localdate(), check_out__isnull=True
    ).exists()


class ValidateQRView(APIView):
    """يتحقق من أن الرمز سارٍ، ويتوقّع الإجراء (دخول/خروج) حسب حالة الموظف الحالي — بدون أي أثر جانبي."""

    permission_classes = [IsEmployeeUser]

    @extend_schema(request=QRTokenInputSerializer, responses={200: ValidateQRResponseSerializer})
    def post(self, request):
        serializer = QRTokenInputSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        action = Action.CHECK_OUT if _has_open_session(request.user) else Action.CHECK_IN
        return Response({"valid": True, "action": action})


def _check_in(user, qr_token: QRToken) -> tuple[Attendance | None, str | None]:
    """ينشئ جلسة جديدة، بشرط مرور الحد الأدنى الزمني منذ آخر خروج لنفس اليوم (إن وجد)
    لمنع تسجيل دخول فوري بعد الخروج بالغلط (مثلاً بسبب مسح مزدوج سريع للرمز).
    القيد الفريد الشرطي على الموديل يمنع أي جلسة مفتوحة ثانية بنفس اليوم حتى تحت التزامن."""
    today = timezone.localdate()
    last_checkout = (
        Attendance.objects.filter(user=user, date=today, check_out__isnull=False)
        .order_by("-check_out")
        .first()
    )
    if last_checkout is not None:
        min_seconds = SystemSettings.get_solo().min_session_duration_seconds
        elapsed = (timezone.now() - last_checkout.check_out).total_seconds()
        if elapsed < min_seconds:
            remaining = int(min_seconds - elapsed)
            return None, f"Please wait {remaining} more second(s) before checking in again."

    try:
        with transaction.atomic():
            attendance = Attendance.objects.create(
                user=user,
                date=today,
                check_in=timezone.now(),
                qr_token=qr_token,
            )
    except IntegrityError:
        return None, "Already checked in today. Check out first before checking in again."
    return attendance, None


def _check_out(user, qr_token: QRToken) -> tuple[Attendance | None, str | None]:
    """يغلق آخر جلسة مفتوحة لهذا الموظف اليوم، بشرط مرور الحد الأدنى الزمني منذ الدخول.
    select_for_update تمنع إغلاق نفس الجلسة مرتين بالتزامن."""
    with transaction.atomic():
        attendance = (
            Attendance.objects.select_for_update()
            .filter(user=user, date=timezone.localdate(), check_out__isnull=True)
            .first()
        )
        if attendance is None:
            return None, "Cannot check out before checking in."

        min_seconds = SystemSettings.get_solo().min_session_duration_seconds
        elapsed = (timezone.now() - attendance.check_in).total_seconds()
        if elapsed < min_seconds:
            remaining = int(min_seconds - elapsed)
            return None, f"Please wait {remaining} more second(s) before checking out."

        attendance.check_out = timezone.now()
        attendance.qr_token = qr_token
        attendance.save()
    return attendance, None


class RecordAttendanceView(APIView):
    """يحفظ الحضور/الانصراف فعليًا — يُستدعى بعد نجاح البصمة محليًا فقط.
    الإجراء (دخول/خروج) يُحدَّد تلقائيًا حسب وجود جلسة مفتوحة للموظف أم لا."""

    permission_classes = [IsEmployeeUser]

    @extend_schema(request=QRTokenInputSerializer, responses={200: AttendanceSerializer})
    def post(self, request):
        serializer = QRTokenInputSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        qr_token = serializer.validated_data["token"]

        if _has_open_session(request.user):
            attendance, error = _check_out(request.user, qr_token)
        else:
            attendance, error = _check_in(request.user, qr_token)

        if error:
            return Response({"detail": error}, status=status.HTTP_400_BAD_REQUEST)
        return Response(AttendanceSerializer(attendance).data)


class MyAttendanceView(generics.ListAPIView):
    serializer_class = AttendanceSerializer
    permission_classes = [IsEmployeeUser]

    def get_queryset(self):
        today = timezone.localdate()
        return Attendance.objects.filter(
            user=self.request.user, date__year=today.year, date__month=today.month
        )


class EmployeeExportAttendanceView(APIView):
    """يصدّر تقرير إكسل لموظف واحد محدَّد بمعرّفه، لأي نطاق تاريخي (من - إلى) يختاره الأدمن."""

    permission_classes = [IsAdminUser]

    @extend_schema(parameters=DATE_RANGE_PARAMETERS, responses={200: OpenApiTypes.BINARY})
    def get(self, request, pk):
        employee = get_object_or_404(User, pk=pk, is_employee=True)
        start_date, end_date = _resolve_date_range(request)

        records = Attendance.objects.filter(
            user=employee, date__gte=start_date, date__lte=end_date
        ).order_by("date")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f"{start_date.isoformat()}_{end_date.isoformat()}"
        sheet.append(["التاريخ", "وقت الحضور", "وقت الانصراف"])

        for record in records:
            sheet.append(
                [
                    record.date.isoformat(),
                    timezone.localtime(record.check_in).strftime("%H:%M:%S")
                    if record.check_in
                    else "",
                    timezone.localtime(record.check_out).strftime("%H:%M:%S")
                    if record.check_out
                    else "",
                ]
            )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f"attachment; filename=attendance_{employee.username}_{start_date}_{end_date}.xlsx"
        )
        workbook.save(response)
        return response


WEEKEND_ISOWEEKDAYS = {6, 7}  # السبت والأحد

USER_FILTERSET_FIELDS = [
    "username",
    "email",
    "first_name",
    "last_name",
    "phone",
    "type",
    "tc",
    "entity",
    "is_admin",
    "is_entry",
    "is_employee",
    "is_regular",
    "is_active",
    "is_first_login",
    "device_id",
]
USER_SEARCH_FIELDS = ["first_name", "last_name", "username", "phone", "tc"]


def _working_days_in_range(year: int, month: int, up_to_day: int) -> int:
    return sum(
        1
        for day in range(1, up_to_day + 1)
        if date(year, month, day).isoweekday() not in WEEKEND_ISOWEEKDAYS
    )


def _resolve_working_days(year: int, month: int) -> int:
    """أيام العمل (باستثناء السبت والأحد) من بداية الشهر حتى اليوم، أو حتى نهاية الشهر لو كان شهرًا سابقًا."""
    today = timezone.localdate()
    days_in_month = calendar.monthrange(year, month)[1]
    last_relevant_day = today.day if (year, month) == (today.year, today.month) else days_in_month
    return _working_days_in_range(year, month, last_relevant_day)


def _count_working_days_between(start_date: date, end_date: date) -> int:
    """أيام العمل (باستثناء السبت والأحد) ضمن نطاق تاريخي حر، محسوبة حتى اليوم فقط إن امتد
    النطاق للمستقبل (لا يمكن اعتبار يوم لم يأتِ بعد غيابًا)."""
    effective_end = min(end_date, timezone.localdate())
    if effective_end < start_date:
        return 0
    return sum(
        1
        for offset in range((effective_end - start_date).days + 1)
        if (start_date + timedelta(days=offset)).isoweekday() not in WEEKEND_ISOWEEKDAYS
    )


def _attendance_stats_for_records(records, *, work_start_time, work_end_time, working_days: int) -> dict:
    """يحسب أيام الدوام/الغياب ودقائق التأخير والانصراف المبكر من مجموعة تسجيلات حضور جاهزة.
    التأخير يُقاس من أول تسجيل حضور باليوم، والانصراف المبكر من آخر تسجيل خروج باليوم."""
    first_check_in_by_day = {}
    last_check_out_by_day = {}
    for record in records:
        first_check_in_by_day.setdefault(record.date, record.check_in)
        if record.check_out:
            current_latest = last_check_out_by_day.get(record.date)
            if current_latest is None or record.check_out > current_latest:
                last_check_out_by_day[record.date] = record.check_out

    daily_late_minutes = {}
    for day, check_in in first_check_in_by_day.items():
        local_check_in = timezone.localtime(check_in)
        scheduled_start = local_check_in.replace(
            hour=work_start_time.hour, minute=work_start_time.minute, second=0, microsecond=0
        )
        late = int((local_check_in - scheduled_start).total_seconds() // 60)
        daily_late_minutes[day] = max(late, 0)

    daily_early_leave_minutes = {}
    for day, check_out in last_check_out_by_day.items():
        local_check_out = timezone.localtime(check_out)
        scheduled_end = local_check_out.replace(
            hour=work_end_time.hour, minute=work_end_time.minute, second=0, microsecond=0
        )
        early = int((scheduled_end - local_check_out).total_seconds() // 60)
        daily_early_leave_minutes[day] = max(early, 0)

    present_days = len(first_check_in_by_day)
    return {
        "present_days": present_days,
        "absent_days": max(working_days - present_days, 0),
        "late_minutes": sum(daily_late_minutes.values()),
        "early_leave_minutes": sum(daily_early_leave_minutes.values()),
        "daily_late_minutes": daily_late_minutes,
        "daily_early_leave_minutes": daily_early_leave_minutes,
    }


def _employee_month_stats(
    employee, year: int, month: int, *, work_start_time, work_end_time, working_days: int
) -> dict:
    records = Attendance.objects.filter(
        user=employee, date__year=year, date__month=month
    ).order_by("check_in")
    return _attendance_stats_for_records(
        records, work_start_time=work_start_time, work_end_time=work_end_time, working_days=working_days
    )


def _employee_range_stats(
    employee, start_date: date, end_date: date, *, work_start_time, work_end_time, working_days: int
) -> dict:
    records = Attendance.objects.filter(
        user=employee, date__gte=start_date, date__lte=end_date
    ).order_by("check_in")
    return _attendance_stats_for_records(
        records, work_start_time=work_start_time, work_end_time=work_end_time, working_days=working_days
    )


EXPORT_SUMMARY_LABELS = {
    "ar": {
        "columns": [
            "اسم الموظف",
            "الرقم الوطني",
            "الجهة",
            "نوع الدوام",
            "أيام الدوام",
            "أيام الغياب",
        ],
        "phone": "الهاتف",
        "not_set": "غير محدد",
        "detail_columns": ["التاريخ", "وقت الحضور", "وقت الانصراف"],
        "back_to_summary": "⬅ العودة للملخص",
    },
    "en": {
        "columns": [
            "Employee name",
            "National ID",
            "Entity",
            "Duty type",
            "Days present",
            "Days absent",
        ],
        "phone": "Phone",
        "not_set": "Not set",
        "detail_columns": ["Date", "Check-in", "Check-out"],
        "back_to_summary": "⬅ Back to summary",
    },
    "tr": {
        "columns": [
            "Çalışan adı",
            "TC kimlik no",
            "Kurum",
            "Çalışma türü",
            "Çalışılan gün",
            "Devamsızlık günü",
        ],
        "phone": "Telefon",
        "not_set": "Belirtilmemiş",
        "detail_columns": ["Tarih", "Giriş saati", "Çıkış saati"],
        "back_to_summary": "⬅ Özete dön",
    },
}

INVALID_SHEET_NAME_CHARS = re.compile(r"[\\/*?:\[\]']")


def _unique_sheet_name(raw_name: str, used_names: set) -> str:
    """يحوّل اسم الموظف إلى اسم ورقة إكسل صالح (حتى 31 محرفًا، بدون رموز محظورة)، ويضمن تفرّده
    ضمن نفس الملف حتى لو تشابهت الأسماء بعد الاقتصاص."""
    cleaned = INVALID_SHEET_NAME_CHARS.sub(" ", raw_name).strip() or "Employee"
    candidate = cleaned[:31]
    suffix = 2
    while candidate.lower() in used_names:
        suffix_text = f" ({suffix})"
        candidate = cleaned[: 31 - len(suffix_text)] + suffix_text
        suffix += 1
    used_names.add(candidate.lower())
    return candidate


HYPERLINK_FONT = Font(color="0563C1", underline="single")

# ترتيب الحقول ثابت دائمًا في ورقة تفاصيل الموظف — تُستخدم أرقام الصفوف هذه لبناء صيغ الربط
# (='اسم الورقة'!خلية) في ورقة الملخص، بحيث يظهر أي تعديل على هذه القيم بالورقة الفرعية تلقائيًا
# في الورقة الرئيسية عند فتح الملف بإكسل.
DETAIL_ROW_NAME = 2
DETAIL_ROW_PHONE = 3
DETAIL_ROW_TC = 4
DETAIL_ROW_ENTITY = 5
DETAIL_ROW_TYPE = 6
DETAIL_ROW_PRESENT_DAYS = 7
DETAIL_ROW_ABSENT_DAYS = 8


def _write_employee_detail_sheet(
    sheet, employee, full_name, stats, labels, start_date, end_date, summary_sheet_title
):
    """يكتب معلومات موظف واحد وتفاصيل حضوره اليومي خلال نطاق تاريخي محدَّد في ورقة إكسل مستقلة
    ضمن نفس ملف الملخص، مع رابط للعودة لورقة الملخص العامة. أرقام صفوف الحقول المشتركة مع الملخص
    ثابتة (انظر DETAIL_ROW_*) لأن ورقة الملخص تربط خلاياها بهذه الخلايا عبر صيغ."""
    back_link = sheet.cell(row=1, column=1, value=labels["back_to_summary"])
    back_link.hyperlink = f"#'{summary_sheet_title}'!A1"
    back_link.font = HYPERLINK_FONT

    sheet.cell(row=DETAIL_ROW_NAME, column=1, value=full_name).font = Font(bold=True, size=14)
    sheet.cell(row=DETAIL_ROW_PHONE, column=1, value=labels["phone"])
    sheet.cell(row=DETAIL_ROW_PHONE, column=2, value=employee.phone or labels["not_set"])
    sheet.cell(row=DETAIL_ROW_TC, column=1, value=labels["columns"][1])
    sheet.cell(
        row=DETAIL_ROW_TC, column=2,
        value=employee.tc if employee.tc is not None else labels["not_set"],
    )
    sheet.cell(row=DETAIL_ROW_ENTITY, column=1, value=labels["columns"][2])
    sheet.cell(
        row=DETAIL_ROW_ENTITY, column=2,
        value=employee.entity if employee.entity is not None else labels["not_set"],
    )
    sheet.cell(row=DETAIL_ROW_TYPE, column=1, value=labels["columns"][3])
    sheet.cell(
        row=DETAIL_ROW_TYPE, column=2,
        value=employee.type if employee.type is not None else labels["not_set"],
    )
    sheet.cell(row=DETAIL_ROW_PRESENT_DAYS, column=1, value=labels["columns"][4])
    sheet.cell(row=DETAIL_ROW_PRESENT_DAYS, column=2, value=stats["present_days"])
    sheet.cell(row=DETAIL_ROW_ABSENT_DAYS, column=1, value=labels["columns"][5])
    sheet.cell(row=DETAIL_ROW_ABSENT_DAYS, column=2, value=stats["absent_days"])

    sheet.append([])
    sheet.append(labels["detail_columns"])

    records = Attendance.objects.filter(
        user=employee, date__gte=start_date, date__lte=end_date
    ).order_by("date")
    for record in records:
        sheet.append(
            [
                record.date.isoformat(),
                timezone.localtime(record.check_in).strftime("%H:%M:%S") if record.check_in else "",
                timezone.localtime(record.check_out).strftime("%H:%M:%S") if record.check_out else "",
            ]
        )


class MonthlyAttendanceSummaryExportView(generics.GenericAPIView):
    """يصدّر تقرير إكسل يلخّص بيانات كل الموظفين الأساسية (الرقم الوطني، الجهة، نوع الدوام) وحضورهم
    خلال نطاق تاريخي محدَّد (من - إلى): أيام دوام، أيام غياب. يُطبَّق نفس فلتر/بحث قائمة الموظفين حتى
    يطابق الملف ما يراه الأدمن على الشاشة، وتُترجَم عناوينه حسب باراميتر lang (ar/en/tr) لتطابق لغة
    الواجهة وقت التصدير."""

    permission_classes = [IsAdminUser]
    queryset = User.objects.filter(is_employee=True).order_by("username")
    filter_backends = [DjangoFilterBackend, drf_filters.SearchFilter]
    filterset_fields = USER_FILTERSET_FIELDS
    search_fields = USER_SEARCH_FIELDS

    @extend_schema(
        parameters=DATE_RANGE_PARAMETERS
        + [OpenApiParameter("lang", OpenApiTypes.STR, OpenApiParameter.QUERY, required=False)],
        responses={200: OpenApiTypes.BINARY},
    )
    def get(self, request):
        start_date, end_date = _resolve_date_range(request)
        settings_obj = SystemSettings.get_solo()
        work_start_time = settings_obj.work_start_time
        work_end_time = settings_obj.work_end_time
        labels = EXPORT_SUMMARY_LABELS.get(request.query_params.get("lang"), EXPORT_SUMMARY_LABELS["ar"])
        working_days = _count_working_days_between(start_date, end_date)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f"{start_date.isoformat()}_{end_date.isoformat()}"
        sheet.append(labels["columns"])

        used_sheet_names = {sheet.title.lower()}

        employees = self.filter_queryset(self.get_queryset())
        for employee in employees:
            stats = _employee_range_stats(
                employee,
                start_date,
                end_date,
                work_start_time=work_start_time,
                work_end_time=work_end_time,
                working_days=working_days,
            )
            full_name = f"{employee.first_name} {employee.last_name}".strip() or employee.username

            detail_sheet_name = _unique_sheet_name(full_name, used_sheet_names)
            detail_sheet = workbook.create_sheet(title=detail_sheet_name)
            _write_employee_detail_sheet(
                detail_sheet, employee, full_name, stats, labels, start_date, end_date, sheet.title
            )

            # صيغ (لا قيم ثابتة) تشير لخلايا ورقة الموظف الفرعية، حتى يظهر أي تعديل عليها هناك
            # تلقائيًا هنا أيضًا عند فتح الملف بإكسل.
            sheet.append(
                [
                    f"='{detail_sheet_name}'!A{DETAIL_ROW_NAME}",
                    f"='{detail_sheet_name}'!B{DETAIL_ROW_TC}",
                    f"='{detail_sheet_name}'!B{DETAIL_ROW_ENTITY}",
                    f"='{detail_sheet_name}'!B{DETAIL_ROW_TYPE}",
                    f"='{detail_sheet_name}'!B{DETAIL_ROW_PRESENT_DAYS}",
                    f"='{detail_sheet_name}'!B{DETAIL_ROW_ABSENT_DAYS}",
                ]
            )
            name_cell = sheet.cell(row=sheet.max_row, column=1)
            name_cell.hyperlink = f"#'{detail_sheet_name}'!A1"
            name_cell.font = HYPERLINK_FONT

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f"attachment; filename=attendance_summary_{start_date}_{end_date}.xlsx"
        )
        workbook.save(response)
        return response


class AdminStatsOverviewView(APIView):
    """لوحة إحصائيات شاملة للأدمن خلال شهر محدَّد: ملخص عام، اتجاه الحضور اليومي، وأكثر الموظفين
    تأخيرًا وغيابًا. لا تُطبَّق عليها فلاتر — تغطي كل الموظفين دائمًا."""

    permission_classes = [IsAdminUser]

    @extend_schema(parameters=YEAR_MONTH_PARAMETERS, responses={200: AdminStatsOverviewSerializer})
    def get(self, request):
        year, month = _resolve_year_month(request)
        settings_obj = SystemSettings.get_solo()
        work_start_time = settings_obj.work_start_time
        work_end_time = settings_obj.work_end_time
        working_days = _resolve_working_days(year, month)

        employees = list(User.objects.filter(is_employee=True).order_by("username"))
        total_employees = len(employees)
        regular_count = sum(1 for employee in employees if employee.is_regular)
        active_count = sum(1 for employee in employees if employee.is_active)

        total_present_days = 0
        total_absent_days = 0
        total_late_minutes = 0
        total_early_leave_minutes = 0
        top_late = []
        top_absent = []
        top_early_leave = []

        for employee in employees:
            stats = _employee_month_stats(
                employee,
                year,
                month,
                work_start_time=work_start_time,
                work_end_time=work_end_time,
                working_days=working_days,
            )
            total_present_days += stats["present_days"]
            total_absent_days += stats["absent_days"]
            total_late_minutes += stats["late_minutes"]
            total_early_leave_minutes += stats["early_leave_minutes"]
            name = f"{employee.first_name} {employee.last_name}".strip() or employee.username
            top_late.append({"id": employee.id, "name": name, "value": stats["late_minutes"]})
            top_absent.append({"id": employee.id, "name": name, "value": stats["absent_days"]})
            top_early_leave.append({"id": employee.id, "name": name, "value": stats["early_leave_minutes"]})

        top_late.sort(key=lambda item: item["value"], reverse=True)
        top_absent.sort(key=lambda item: item["value"], reverse=True)
        top_early_leave.sort(key=lambda item: item["value"], reverse=True)

        daily_trend = list(
            Attendance.objects.filter(user__is_employee=True, date__year=year, date__month=month)
            .values("date")
            .annotate(present_count=Count("user", distinct=True))
            .order_by("date")
        )

        possible_present_days = total_employees * working_days
        attendance_rate = (
            round((total_present_days / possible_present_days) * 100, 1) if possible_present_days else 0.0
        )

        data = {
            "year": year,
            "month": month,
            "working_days": working_days,
            "total_employees": total_employees,
            "regular_count": regular_count,
            "irregular_count": total_employees - regular_count,
            "active_count": active_count,
            "suspended_count": total_employees - active_count,
            "entry_account_count": User.objects.filter(is_entry=True).count(),
            "attendance_rate": attendance_rate,
            "total_present_days": total_present_days,
            "total_absent_days": total_absent_days,
            "total_late_minutes": total_late_minutes,
            "total_early_leave_minutes": total_early_leave_minutes,
            "daily_trend": daily_trend,
            "top_late": top_late[:5],
            "top_absent": top_absent[:5],
            "top_early_leave": top_early_leave[:5],
        }
        return Response(AdminStatsOverviewSerializer(data).data)


class EmployeeStatsView(APIView):
    """إحصائيات الحضور الشخصية للموظف الحالي (وحده) خلال شهر محدَّد."""

    permission_classes = [IsEmployeeUser]

    @extend_schema(parameters=YEAR_MONTH_PARAMETERS, responses={200: EmployeeStatsSerializer})
    def get(self, request):
        year, month = _resolve_year_month(request)
        settings_obj = SystemSettings.get_solo()
        working_days = _resolve_working_days(year, month)

        stats = _employee_month_stats(
            request.user,
            year,
            month,
            work_start_time=settings_obj.work_start_time,
            work_end_time=settings_obj.work_end_time,
            working_days=working_days,
        )
        present_days = stats["present_days"]
        on_time_days = sum(
            1
            for day, late_minutes in stats["daily_late_minutes"].items()
            if late_minutes == 0 and stats["daily_early_leave_minutes"].get(day, 0) == 0
        )
        on_time_rate = round((on_time_days / present_days) * 100, 1) if present_days else 0.0

        data = {
            "year": year,
            "month": month,
            "working_days": working_days,
            "present_days": present_days,
            "absent_days": stats["absent_days"],
            "late_minutes": stats["late_minutes"],
            "early_leave_minutes": stats["early_leave_minutes"],
            "on_time_rate": on_time_rate,
            "daily_late_minutes": [
                {"date": day, "late_minutes": minutes}
                for day, minutes in sorted(stats["daily_late_minutes"].items())
            ],
            "daily_early_leave_minutes": [
                {"date": day, "early_leave_minutes": minutes}
                for day, minutes in sorted(stats["daily_early_leave_minutes"].items())
            ],
        }
        return Response(EmployeeStatsSerializer(data).data)


class EmployeeListView(generics.ListAPIView):
    queryset = User.objects.filter(is_employee=True).order_by("username")
    serializer_class = UserSerializer
    permission_classes = [IsAdminUser]
    filter_backends = [DjangoFilterBackend, drf_filters.SearchFilter]
    filterset_fields = USER_FILTERSET_FIELDS
    search_fields = USER_SEARCH_FIELDS


class EntryUserListView(generics.ListAPIView):
    queryset = User.objects.filter(is_entry=True).order_by("username")
    serializer_class = UserSerializer
    permission_classes = [IsAdminUser]
    filter_backends = [DjangoFilterBackend, drf_filters.SearchFilter]
    filterset_fields = USER_FILTERSET_FIELDS
    search_fields = USER_SEARCH_FIELDS


class EmployeeDetailView(APIView):
    permission_classes = [IsAdminUser]

    @extend_schema(parameters=DATE_RANGE_PARAMETERS, responses={200: EmployeeAttendanceSerializer})
    def get(self, request, pk):
        employee = get_object_or_404(User, pk=pk, is_employee=True)
        start_date, end_date = _resolve_date_range(request)
        attendance = Attendance.objects.filter(
            user=employee, date__gte=start_date, date__lte=end_date
        ).order_by("date")

        settings_obj = SystemSettings.get_solo()
        working_days = _count_working_days_between(start_date, end_date)
        stats = _employee_range_stats(
            employee,
            start_date,
            end_date,
            work_start_time=settings_obj.work_start_time,
            work_end_time=settings_obj.work_end_time,
            working_days=working_days,
        )

        data = UserSerializer(employee).data
        data["attendance"] = AttendanceSerializer(attendance, many=True).data
        data["present_days"] = stats["present_days"]
        data["absent_days"] = stats["absent_days"]
        data["late_minutes"] = stats["late_minutes"]
        data["early_leave_minutes"] = stats["early_leave_minutes"]
        return Response(data)


class SystemSettingsView(APIView):
    """يدير الأدمن منها مدة صلاحية رمز QR — نفس المدة تُستخدم كفاصل التحديث التلقائي بالفرونت."""

    permission_classes = [IsAdminUser]

    @extend_schema(responses={200: SystemSettingsSerializer})
    def get(self, request):
        return Response(SystemSettingsSerializer(SystemSettings.get_solo()).data)

    @extend_schema(request=SystemSettingsSerializer, responses={200: SystemSettingsSerializer})
    def patch(self, request):
        serializer = SystemSettingsSerializer(
            SystemSettings.get_solo(), data=request.data, partial=True
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)
